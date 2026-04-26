"""
ai/engine.py  — 100% IN-HOUSE AI ENGINE
=========================================
Zero external API calls. Zero Claude/OpenAI/Gemini dependencies.
All intelligence built using: pdfplumber + openpyxl + regex + rapidfuzz

Handles:
  1. BankStatementParser  — SBI, HDFC, ICICI, Axis, Kotak, Yes Bank,
                            IDFC First, Federal, PNB, BOI, Canara, UCO
                            CSV / Excel / PDF / Plain text
  2. InvoiceParser        — GST tax invoices, purchase bills, receipts
  3. VoucherParser        — Auto-detect voucher type from narration
  4. TransactionClassifier— 400+ rules for Indian business transactions
  5. ConfidenceScorer     — Assigns confidence score per classification

Author: Finix AI · In-house · No API keys required
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

# ── Optional deps (graceful fallback) ────────────────────────────────────────
try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    pdfplumber = None
    PDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    openpyxl = None
    EXCEL_AVAILABLE = False

try:
    from rapidfuzz import fuzz
    FUZZY_AVAILABLE = True
except ImportError:
    fuzz = None
    FUZZY_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════════════════
# DATA CLASSES
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ParsedTransaction:
    txn_date:   str             # YYYY-MM-DD
    narration:  str
    amount:     float
    txn_type:   str             # "credit" | "debit"
    balance:    float = 0.0
    reference:  str   = ""
    value_date: str   = ""

@dataclass
class ClassifiedTransaction(ParsedTransaction):
    account:    str   = "Miscellaneous Expense"
    confidence: float = 0.65
    method:     str   = "rule"
    status:     str   = "unmatched"
    id:         str   = ""

@dataclass
class ParsedInvoice:
    invoice_no:   str   = ""
    party_name:   str   = ""
    party_gstin:  str   = ""
    invoice_date: str   = ""
    narration:    str   = ""
    hsn_sac:      str   = ""
    amount:       float = 0.0
    cgst:         float = 0.0
    sgst:         float = 0.0
    igst:         float = 0.0
    total:        float = 0.0
    inv_type:     str   = "purchase"   # "sales" | "purchase"

@dataclass
class ParsedVoucher:
    voucher_type: str   = "journal"    # journal|payment|receipt|contra|sales|purchase
    date:         str   = ""
    narration:    str   = ""
    reference:    str   = ""
    party:        str   = ""
    amount:       float = 0.0
    debit_account:  str = ""
    credit_account: str = ""
    confidence:   float = 0.70


# ═══════════════════════════════════════════════════════════════════════════════
# 1. BANK STATEMENT PARSER  (100% in-house)
# ═══════════════════════════════════════════════════════════════════════════════

class BankStatementParser:
    """
    Parses bank statements from any Indian bank.
    NO external API. Uses pdfplumber for PDF + regex for text.
    Handles SBI, HDFC, ICICI, Axis, Kotak, Yes Bank, IDFC, Federal,
    PNB, BOI, Canara, UCO — CSV / Excel / PDF / Text.
    """

    DATE_FORMATS = [
        "%d %b %Y", "%d-%b-%Y", "%d/%b/%Y",
        "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
        "%Y-%m-%d", "%m/%d/%Y",
        "%d %b %y", "%d-%b-%y",
        "%d-%m-%y", "%d/%m/%y",
    ]

    AMOUNT_RE = re.compile(r"[\d,]+\.?\d*")

    # Column name signatures per bank
    BANK_SIGNATURES = {
        "sbi": {
            "date":      ["txn date", "date"],
            "narration": ["description", "particulars", "narration"],
            "debit":     ["debit"],
            "credit":    ["credit"],
            "balance":   ["balance"],
            "reference": ["ref no", "cheque no", "reference"],
        },
        "hdfc": {
            "date":      ["date"],
            "narration": ["narration", "description"],
            "debit":     ["withdrawal", "debit"],
            "credit":    ["deposit", "credit"],
            "balance":   ["closing balance", "balance"],
            "reference": ["chq/ref", "ref no", "reference"],
        },
        "icici": {
            "date":      ["transaction date", "value date", "date"],
            "narration": ["transaction remarks", "description", "particulars"],
            "debit":     ["withdrawal", "debit", "dr"],
            "credit":    ["deposit", "credit", "cr"],
            "balance":   ["balance"],
            "reference": ["reference no", "reference"],
        },
        "axis": {
            "date":      ["tran date", "transaction date", "date"],
            "narration": ["particulars", "description", "narration"],
            "debit":     ["debit", "dr"],
            "credit":    ["credit", "cr"],
            "balance":   ["balance"],
            "reference": ["chq no", "reference no"],
        },
        "kotak": {
            "date":      ["transaction date", "date"],
            "narration": ["description", "narration"],
            "debit":     ["debit"],
            "credit":    ["credit"],
            "balance":   ["balance"],
            "reference": ["reference no"],
        },
        "generic": {
            "date":      ["date", "txn date", "transaction date", "value date",
                          "posting date", "trans date"],
            "narration": ["narration", "description", "particulars", "remarks",
                          "details", "transaction remarks", "memo"],
            "debit":     ["debit", "withdrawal", "dr", "debit amount",
                          "withdrawal amt", "paid out", "withdrawals"],
            "credit":    ["credit", "deposit", "cr", "credit amount",
                          "deposit amt", "paid in", "received", "deposits"],
            "balance":   ["balance", "closing balance", "running balance",
                          "available balance", "bal"],
            "reference": ["reference", "ref no", "chq/ref", "cheque no",
                          "utr", "transaction id", "trans id", "ref no./cheque no."],
        },
    }

    def parse(self, content: bytes, filename: str) -> list[ParsedTransaction]:
        """
        Auto-detect file type and parse bank statement.
        Returns list of ParsedTransaction sorted by date ascending.
        """
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else "txt"

        if ext == "pdf":
            txns = self._parse_pdf(content)
        elif ext in ("xlsx", "xls"):
            txns = self._parse_excel(content, ext)
        elif ext == "csv":
            txns = self._parse_csv(content)
        else:
            txns = self._parse_csv(content)
            if not txns:
                txns = self._parse_text(content.decode("utf-8-sig", errors="replace"))

        txns.sort(key=lambda t: t.txn_date)
        return txns

    # ── PDF Parser ────────────────────────────────────────────────────────────

    def _parse_pdf(self, content: bytes) -> list[ParsedTransaction]:
        if not PDF_AVAILABLE:
            raise RuntimeError(
                "pdfplumber is not installed on this server. "
                "Make sure pdfplumber==0.11.1 is in requirements-render.txt and redeploy."
            )
        transactions: list[ParsedTransaction] = []

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                # Strategy 1: structured table extraction
                tables = page.extract_tables()
                for table in tables:
                    txns = self._parse_table(table)
                    transactions.extend(txns)

                # Strategy 2: if no table or empty, try text-based parsers
                if not tables or not transactions:
                    text = page.extract_text() or ""
                    # Try SBI-specific parser first (dual-date format)
                    txns = self._parse_sbi_text(text)
                    if not txns:
                        # Try HDFC/ICICI text format
                        txns = self._parse_text_lines(text.split("\n"))
                    transactions.extend(txns)

        # Deduplicate
        seen = set()
        unique = []
        for t in transactions:
            key = (t.txn_date, t.amount, t.narration[:40])
            if key not in seen:
                seen.add(key)
                unique.append(t)

        return unique

    def _parse_sbi_text(self, text: str) -> list[ParsedTransaction]:
        """
        SBI-specific text parser.
        SBI PDF format: each transaction row has TWO dates followed by
        narration, reference, branch code, debit/credit, balance.
        Lines wrap — collect all lines until next date-pair row.

        Example:
          4 Apr 2025 4 Apr 2025 BY TRANSFER-UPI/CR/...  1,770.00  28,672.17
        """
        transactions: list[ParsedTransaction] = []
        lines = [l.rstrip() for l in text.split("\n") if l.strip()]

        # SBI rows start with two dates like "4 Apr 2025 4 Apr 2025"
        DATE_RE = re.compile(r"^(\d{1,2}\s+\w{3}\s+\d{4})\s+(\d{1,2}\s+\w{3}\s+\d{4})")
        AMT_RE  = re.compile(r"([\d,]+\.\d{2})")

        i = 0
        while i < len(lines):
            m = DATE_RE.match(lines[i])
            if not m:
                i += 1
                continue

            txn_date_raw = m.group(1)
            txn_date     = self._parse_date(txn_date_raw)
            if not txn_date:
                i += 1
                continue

            # Collect continuation lines until next date-pair row (max 10 lines)
            block = [lines[i]]
            j = i + 1
            while j < len(lines) and j < i + 10 and not DATE_RE.match(lines[j]):
                block.append(lines[j])
                j += 1

            block_text = " ".join(block)

            # Extract all currency-format numbers
            amounts = []
            for a in AMT_RE.findall(block_text):
                try:
                    amounts.append(float(a.replace(",", "")))
                except ValueError:
                    pass

            if len(amounts) < 2:
                i = j
                continue

            # Last = balance, second-to-last = transaction amount
            balance    = amounts[-1]
            txn_amount = amounts[-2]

            if txn_amount == 0:
                i = j
                continue

            # Determine credit/debit from narration keywords
            bu = block_text.upper()
            if "BY DEBIT CARD" in bu or "OTHPG" in bu or "ATM WDL" in bu:
                txn_type = "debit"
            elif any(kw in bu for kw in [
                "BY TRANSFER", "BY CLEARING", "UPI/CR", "BULK POSTING",
                "CASH CREDIT", "CHEQUE DEPOSIT",
            ]):
                txn_type = "credit"
            elif any(kw in bu for kw in [
                "TO TRANSFER", "TO CLEARING", "ATM", "IMPS/",
                "TRANSFER TO", "OUT-CHQ", "CASH CHEQUE",
            ]):
                txn_type = "debit"
            else:
                # Balance-movement heuristic
                txn_type = "credit" if len(amounts) >= 3 and amounts[-1] > amounts[-3] else "debit"

            # Build clean narration
            narration_part = re.sub(r"\d{1,2}\s+\w{3}\s+\d{4}", "", block_text)
            narration_part = AMT_RE.sub("", narration_part)
            narration_part = re.sub(r"\s+", " ", narration_part).strip()
            narration      = self._clean_narration(narration_part or "Bank Transaction")

            transactions.append(ParsedTransaction(
                txn_date  = txn_date,
                narration = narration,
                amount    = txn_amount,
                txn_type  = txn_type,
                balance   = balance,
            ))
            i = j

        return transactions

    def _parse_table(self, table: list[list]) -> list[ParsedTransaction]:
        """Parse a pdfplumber table into transactions."""
        if not table or len(table) < 2:
            return []

        header_row = None
        header_idx = 0
        for i, row in enumerate(table):
            if row and any(str(c or "").strip() for c in row):
                header_row = [str(c or "").strip().lower() for c in row]
                header_idx = i
                break

        if not header_row:
            return []

        col = self._map_columns(header_row, "generic")
        if col["date"] is None:
            return []

        transactions = []
        for row in table[header_idx + 1:]:
            if not row or len(row) < 2:
                continue
            t = self._row_to_transaction(row, col)
            if t:
                transactions.append(t)

        return transactions

    def _map_columns(self, headers: list[str], bank: str = "generic") -> dict:
        """Map column names to indices."""
        sig = self.BANK_SIGNATURES.get(bank, self.BANK_SIGNATURES["generic"])
        result = {
            "date": None, "narration": None, "debit": None,
            "credit": None, "balance": None, "reference": None,
        }
        for i, h in enumerate(headers):
            h_clean = h.lower().strip()
            for field_name, keywords in sig.items():
                if result[field_name] is None:
                    for kw in keywords:
                        if kw in h_clean:
                            result[field_name] = i
                            break

        # Fallback to generic
        if result["date"] is None or result["narration"] is None:
            generic = self.BANK_SIGNATURES["generic"]
            for i, h in enumerate(headers):
                h_clean = h.lower().strip()
                for field_name, keywords in generic.items():
                    if result[field_name] is None:
                        for kw in keywords:
                            if kw in h_clean:
                                result[field_name] = i
                                break
        return result

    def _row_to_transaction(self, row: list, col: dict) -> Optional[ParsedTransaction]:
        """Convert a table row to ParsedTransaction."""
        def get(idx):
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        raw_date = get(col["date"])
        txn_date = self._parse_date(raw_date)
        if not txn_date:
            return None

        narration = get(col["narration"]).replace("\n", " ")
        if not narration or narration.lower() in ("narration", "description", "particulars"):
            return None

        debit  = self._parse_amount(get(col["debit"]))
        credit = self._parse_amount(get(col["credit"]))

        if debit > 0:
            amount, txn_type = debit, "debit"
        elif credit > 0:
            amount, txn_type = credit, "credit"
        else:
            return None

        balance   = self._parse_amount(get(col["balance"]))
        reference = get(col["reference"])

        return ParsedTransaction(
            txn_date  = txn_date,
            narration = self._clean_narration(narration),
            amount    = amount,
            txn_type  = txn_type,
            balance   = balance,
            reference = reference,
        )

    # ── CSV Parser ────────────────────────────────────────────────────────────

    def _parse_csv(self, content: bytes) -> list[ParsedTransaction]:
        """Parse CSV bank statement — handles any column layout."""
        for encoding in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = content.decode("utf-8", errors="replace")

        reader = csv.reader(io.StringIO(text))
        rows   = list(reader)

        header_idx = 0
        for i, row in enumerate(rows):
            if sum(1 for c in row if c.strip()) >= 3:
                header_idx = i
                break

        header = [c.strip().lower() for c in rows[header_idx]]
        col    = self._map_columns(header, "generic")

        if col["date"] is None:
            return []

        transactions = []
        for row in rows[header_idx + 1:]:
            if not row or sum(1 for c in row if c.strip()) < 2:
                continue
            t = self._row_to_transaction(row, col)
            if t:
                transactions.append(t)

        return transactions

    # ── Excel Parser ──────────────────────────────────────────────────────────

    def _parse_excel(self, content: bytes, ext: str) -> list[ParsedTransaction]:
        """Parse Excel bank statement."""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl is not installed.")

        if ext == "xls":
            try:
                import xlrd
                wb   = xlrd.open_workbook(file_contents=content)
                ws   = wb.sheet_by_index(0)
                rows = [[str(ws.cell_value(r, c) or "") for c in range(ws.ncols)]
                        for r in range(ws.nrows)]
            except Exception:
                return []
        else:
            wb   = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws   = wb.active
            rows = [[str(c.value or "") for c in row] for row in ws.iter_rows()]

        if not rows:
            return []

        header_idx = 0
        for i, row in enumerate(rows[:20]):
            if sum(1 for c in row if c.strip()) >= 3:
                header_idx = i
                break

        header = [c.strip().lower() for c in rows[header_idx]]
        col    = self._map_columns(header, "generic")

        if col["date"] is None:
            return []

        transactions = []
        for row in rows[header_idx + 1:]:
            if not row or sum(1 for c in row if c.strip()) < 2:
                continue
            t = self._row_to_transaction(row, col)
            if t:
                transactions.append(t)

        return transactions

    # ── Text / Fallback Parser ────────────────────────────────────────────────

    def _parse_text(self, text: str) -> list[ParsedTransaction]:
        return self._parse_text_lines(text.split("\n"))

    def _parse_text_lines(self, lines: list[str]) -> list[ParsedTransaction]:
        """Regex-based line parser for raw text bank statements (HDFC, ICICI, etc.)."""
        transactions = []
        PATTERN = re.compile(
            r"(\d{1,2}[\s/-](?:\w{3}|\d{2})[\s/-]\d{2,4})"
            r"\s+(.+?)\s+"
            r"([\d,]+\.?\d*)\s*"
            r"(?:([\d,]+\.?\d*)\s*)?"
            r"([\d,]+\.?\d*)\s*$",
            re.MULTILINE
        )

        for line in lines:
            line = line.strip()
            if len(line) < 15:
                continue

            m = PATTERN.search(line)
            if not m:
                continue

            txn_date = self._parse_date(m.group(1).strip())
            if not txn_date:
                continue

            narration = self._clean_narration(m.group(2).strip())
            if not narration:
                continue

            amt1    = self._parse_amount(m.group(3) or "")
            amt2    = self._parse_amount(m.group(4) or "")
            balance = self._parse_amount(m.group(5) or "")

            if amt1 > 0 and amt2 > 0:
                if self._is_credit_narration(narration):
                    amount, txn_type = amt2 if amt2 > 0 else amt1, "credit"
                else:
                    amount, txn_type = amt1, "debit"
            elif amt1 > 0:
                txn_type = "credit" if self._is_credit_narration(narration) else "debit"
                amount   = amt1
            else:
                continue

            transactions.append(ParsedTransaction(
                txn_date  = txn_date,
                narration = narration,
                amount    = amount,
                txn_type  = txn_type,
                balance   = balance,
            ))

        return transactions

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _parse_date(self, raw: str) -> Optional[str]:
        if not raw:
            return None
        raw = raw.strip().replace("  ", " ")
        for fmt in self.DATE_FORMATS:
            try:
                return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        # Numeric only: 20250404
        digits = re.sub(r"\D", "", raw)
        if len(digits) == 8:
            for fmt in ("%d%m%Y", "%Y%m%d"):
                try:
                    return datetime.strptime(digits, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    pass
        return None

    def _parse_amount(self, raw: str) -> float:
        if not raw:
            return 0.0
        cleaned = re.sub(r"[₹$€£\s]", "", raw).replace(",", "")
        m = re.search(r"\d+\.?\d*", cleaned)
        if not m:
            return 0.0
        try:
            val = float(m.group())
            return val if val > 0 else 0.0
        except ValueError:
            return 0.0

    def _clean_narration(self, text: str) -> str:
        text = re.sub(r"[\x00-\x1f\x7f]", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        text = re.sub(r"\bNaN\b", "", text)
        return text[:200]

    def _is_credit_narration(self, narration: str) -> bool:
        n = narration.upper()
        credit_kw = [
            "BY TRANSFER", "BY CLEARING", "CREDIT", "RECEIVED", "RECEIPT",
            "REFUND", "REVERSAL", "RETURN", "DIVIDEND", "INTEREST CREDIT",
            "UPI/CR", "NEFT/CR", "IMPS/CR", "CR/", "/CR/", "BULK POSTING",
        ]
        debit_kw = [
            "TO TRANSFER", "TO CLEARING", "DEBIT", "PAID", "PAYMENT",
            "WITHDRAWAL", "ATM WDL", "PURCHASE", "UPI/DR", "NEFT/DR",
        ]
        credit_score = sum(1 for kw in credit_kw if kw in n)
        debit_score  = sum(1 for kw in debit_kw  if kw in n)
        return credit_score > debit_score


# ═══════════════════════════════════════════════════════════════════════════════
# 2. TRANSACTION CLASSIFIER  (400+ in-house rules)
# ═══════════════════════════════════════════════════════════════════════════════

class TransactionClassifier:
    """
    Classifies Indian bank transaction narrations into accounting ledger accounts.
    400+ pattern rules covering common Indian business transactions.
    NO external AI API. Uses keyword matching + fuzzy fallback.
    """

    # ── Master rule table: (keywords, account_name, confidence) ──────────────
    RULES: list[tuple[list[str], str, float]] = [

        # Salaries & Payroll
        (["salary", "sal/", "salaries", "payroll", "wages", "pay slip",
          "staff payment", "employee payment", "staff sal", "monthly sal",
          "empl pay", "/sal/", "stipend", "remuneration"],
         "Salaries & Wages", 0.95),

        # Rent
        (["rent ", "rental", "lease rent", "office rent", "shop rent",
          "premises rent", "building rent", "flat rent", "property rent",
          "monthly rent", "godown rent"],
         "Rent", 0.93),

        # Electricity & Utilities
        (["electricity", "bescom", "tata power", "msedcl", "tneb", "cesc",
          "wbsedcl", "adani electric", "torrent power", "bses", "dvvnl",
          "uppcl", "pspcl", "kseb", "mepdcl", "electric bill",
          "power bill", "hathway", "act fibernet", "d2h", "dish tv",
          "tata sky", "airtel dth"],
         "Electricity & Utilities", 0.93),

        # Internet & Telephone
        (["jio ", "airtel", "bsnl", "vodafone", "vi ", "idea ", "mtnl",
          "reliance jio", "broadband", "internet", "telecom", "telephone",
          "mobile recharge", "postpaid", "wi-fi", "wifi"],
         "Electricity & Utilities", 0.90),

        # GST Payments
        (["gst payment", "gstpay", "kotakgstpay", "gstin", "gst challan",
          "goods and services tax", "gst return", "gst filing",
          "cmp gst", "gstn", "kotak gst", "axis gst", "hdfc gst",
          "cgst payment", "sgst payment", "igst payment",
          "goods and services taxnew delhi"],
         "GST Payment", 0.97),

        # TDS & Income Tax
        (["tds payment", "income tax", "advance tax", "self assessment",
          "it dept", "itdtax", "tax refund", "itr", "tds challan",
          "tan payment", "income-tax", "tin nsdl", "oltas",
          "tax deducted", "tds return"],
         "TDS Payment", 0.97),

        # PF & ESIC
        (["epfo", "pf payment", "provident fund", "esic payment",
          "employees pf", "pf challan", "esic challan", "esi payment",
          "nps contribution", "gratuity", "pf deposit"],
         "Salaries & Wages", 0.94),

        # Bank Charges & DSC
        (["bank charge", "service charge", "annual fee", "neft charge",
          "rtgs charge", "sms charge", "processing fee", "bank fee",
          "amc charge", "demat charge", "locker charge", "card fee",
          "maintenance charge", "acct keeping", "a/c keeping",
          "pantagon sign", "agon sign", "dsc", "digital sign",
          "sign securi", "certificate", "controller general",
          "cheque return", "chq return", "out-chq", "bounce charge",
          "ecs return", "nach return", "surcharge", "gst on charges",
          "atm card amc"],
         "Bank Charges", 0.93),

        # Software & Subscriptions
        (["amazon web services", "aws", "google cloud", "azure", "microsoft",
          "adobe", "atlassian", "notion", "slack", "zoom", "github",
          "godaddy", "namecheap", "hostgator", "bluehost", "shopify",
          "tally", "busy software", "zoho", "freshbooks", "quickbooks",
          "saas", "subscription", "software", "eloquent info",
          "hathway sales", "claude.ai"],
         "Software Subscriptions", 0.92),

        # Advertising & Marketing
        (["google ads", "facebook", "meta ", "instagram", "youtube ads",
          "linkedin", "twitter", "advertising", "marketing", "promotion",
          "ad spend", "campaign", "pamphlet", "banner", "www.facebook",
          "apnaco", "razorpay marketing", "digital marketing",
          "seo ", "smm ", "social media"],
         "Advertising & Marketing", 0.93),

        # Professional Fees
        (["ca fees", "audit fees", "legal fees", "advocate", "consultant",
          "professional fee", "advisory", "chartered accountant",
          "legal charges", "statutory audit", "tax consultant",
          "manthan desai", "gaya business service", "sublime consultancy",
          "manthan", "deasi", "retainer", "filing charges"],
         "Professional Fees", 0.92),

        # Travel & Conveyance
        (["ola ", "uber", "rapido", "makemytrip", "goibibo", "yatra",
          "irctc", "indigo", "spicejet", "air india", "flight ticket",
          "hotel", "travel", "conveyance", "cab", "taxi", "bus ticket",
          "train ticket", "boarding pass", "lodge", "lodge charges",
          "ntrp", "railway", "petrol", "fuel", "diesel"],
         "Travel & Conveyance", 0.91),

        # Loan Repayment
        (["emi", "loan repayment", "loan emi", "mortgage", "home loan",
          "vehicle loan", "term loan", "hdfc loan", "icici loan",
          "sbi loan", "axis loan", "bajaj finance", "fullerton",
          "muthoot", "manappuram", "lic premium", "insurance premium",
          "loan payment", "principal repayment", "od repayment"],
         "Loan Repayment", 0.94),

        # Insurance
        (["insurance", "lic ", "bajaj allianz", "new india",
          "united india", "national insurance", "oriental insurance",
          "star health", "max bupa", "hdfc ergo", "icici lombard",
          "tata aig", "premium", "policy renewal", "mediclaim",
          "health insurance", "fire insurance", "insure"],
         "Insurance Premium", 0.93),

        # ATM Cash Withdrawals
        (["atm wdl", "atm cash", "cash wdl", "atm withdrawal",
          "cash withdrawal", "atm ", "cash at atm"],
         "ATM Cash Withdrawal", 0.98),

        # Office Supplies & Printing
        (["stationery", "office supplies", "amazon.in", "flipkart",
          "toner", "cartridge", "printer", "photocopy", "binding",
          "printing", "office depot", "paper ", "ink "],
         "Office Supplies", 0.88),

        # Interest Income
        (["interest credit", "int credit", "fd interest", "saving interest",
          "interest on fd", "interest received", "bank interest credit",
          "sbi interest", "hdfc interest", "recurring deposit interest"],
         "Interest Income", 0.94),

        # Interest Expense
        (["interest debited", "interest charged", "od interest",
          "cc interest", "overdue interest", "penal interest",
          "bank interest debit", "finance charge"],
         "Interest Expense", 0.93),

        # Sales / Revenue (client payments)
        (["sales payment", "invoice payment", "payment received",
          "client payment", "customer payment", "receivable",
          "spectevo", "q1account", "q2account", "invoice no",
          "against invoice", "bill no"],
         "Sales Revenue", 0.88),

        # Purchases / Materials
        (["purchase", "supplier payment", "vendor payment", "material",
          "raw material", "stock purchase", "goods purchase",
          "kdk software", "k d k software"],
         "Purchase/Materials", 0.87),

        # Repairs & Maintenance
        (["repair", "maintenance", "amc", "service charge for machine",
          "pest control", "plumbing", "electrical repair",
          "generator", "ac service", "vehicle service", "servicing"],
         "Repairs & Maintenance", 0.89),

        # Medical / Health
        (["hospital", "medical", "doctor", "pharmacy", "medicine",
          "clinic", "health", "apollo", "fortis", "max hospital"],
         "Medical Expenses", 0.87),

        # Food & Canteen
        (["canteen", "food", "lunch", "dinner", "swiggy", "zomato",
          "restaurant", "meal", "snacks"],
         "Staff Welfare", 0.85),

        # Govt Payments / MCA
        (["mca ", "roc fee", "mca21", "company registration",
          "trademark", "patent", "copyright", "govt fee",
          "court fee", "stamp duty"],
         "Government Fees", 0.93),

        # NEFT/UPI generic credits (classify as misc income)
        (["by transfer", "upi/cr", "neft*", "by clearing",
          "imps/cr", "bulk posting"],
         "Miscellaneous Income", 0.72),

        # NEFT/UPI generic debits (classify as misc expense)
        (["to transfer", "upi/dr", "transfer to",
          "imps/dr", "to clearing"],
         "Miscellaneous Expense", 0.72),

        # Cheque related
        (["cheque deposit", "chq deposit",
          "cheque return", "out-chq return", "chq return"],
         "Miscellaneous Income", 0.70),

        # Cash deposits
        (["cash deposit", "cash chq", "cash credit", "counter deposit"],
         "Miscellaneous Income", 0.75),
    ]

    # ── Vendor-specific exact mappings ────────────────────────────────────────
    VENDOR_MAP: dict[str, tuple[str, float]] = {
        # DSC / Digital Signature vendors
        "pantagon":        ("Bank Charges", 0.97),
        "agon sign":       ("Bank Charges", 0.97),
        "sign securi":     ("Bank Charges", 0.97),
        "ntrp":            ("Travel & Conveyance", 0.92),
        # IT Refund
        "itdtax refund":   ("TDS Payment", 0.99),
        "cmp itro":        ("TDS Payment", 0.99),
        # CA / Bookkeeping services
        "manthan desai":   ("Professional Fees", 0.98),
        "gaya business":   ("Professional Fees", 0.98),
        "sublime":         ("Professional Fees", 0.95),
        # Software
        "kdk software":    ("Software Subscriptions", 0.96),
        "claude.ai":       ("Software Subscriptions", 0.99),
        "eloquent info":   ("Software Subscriptions", 0.96),
        # Known clients (SPECTEVO/PRODIGIST accounts)
        "spectevo":        ("Sales Revenue", 0.97),
        "tapipe fintech":  ("Sales Revenue", 0.95),
        "bookends hosp":   ("Sales Revenue", 0.95),
        "synbus tech":     ("Sales Revenue", 0.95),
        "vebnor fashion":  ("Sales Revenue", 0.95),
        "imon technolog":  ("Sales Revenue", 0.95),
        "shanta g foods":  ("Sales Revenue", 0.95),
        "nurene life":     ("Sales Revenue", 0.95),
        "shiv textile":    ("Sales Revenue", 0.95),
        "giga corporat":   ("Sales Revenue", 0.95),
        "vs internation":  ("Sales Revenue", 0.95),
        "rm supplier":     ("Sales Revenue", 0.95),
        "new india fire":  ("Insurance Premium", 0.97),
        "new india tech":  ("Sales Revenue", 0.95),
        "h and h health":  ("Sales Revenue", 0.95),
        "dhananjay crea":  ("Sales Revenue", 0.95),
        "playfair sport":  ("Sales Revenue", 0.95),
        "coach for life":  ("Sales Revenue", 0.95),
        "yashasvee spir":  ("Sales Revenue", 0.95),
        "belizzi":         ("Sales Revenue", 0.95),
        "connectify":      ("Sales Revenue", 0.95),
        "the moov":        ("Sales Revenue", 0.95),
        "pal by rahul":    ("Sales Revenue", 0.95),
        "sigdi re":        ("Sales Revenue", 0.95),
        # Facebook Ads
        "facebook com ads":("Advertising & Marketing", 0.99),
        "www facebook":    ("Advertising & Marketing", 0.99),
        "www.facebook":    ("Advertising & Marketing", 0.99),
        # GST
        "kotakgstpay":     ("GST Payment", 0.99),
        "goods and services taxnew": ("GST Payment", 0.99),
    }

    def classify(self, narration: str) -> tuple[str, float]:
        """
        Classify narration → (account_name, confidence).
        Order: vendor map → rule matching → fuzzy fallback.
        """
        n = narration.upper().strip()

        # 1. Exact vendor match (highest priority)
        for vendor, (account, conf) in self.VENDOR_MAP.items():
            if vendor.upper() in n:
                return account, conf

        # 2. Rule-based keyword match
        for keywords, account, confidence in self.RULES:
            for kw in keywords:
                if kw.upper() in n:
                    boost = 0.03 if len(kw) > 6 else 0.0
                    return account, min(confidence + boost, 0.99)

        # 3. Fuzzy fallback
        if FUZZY_AVAILABLE:
            result = self._fuzzy_classify(narration)
            if result:
                return result

        return "Miscellaneous Expense", 0.60

    def _fuzzy_classify(self, narration: str) -> Optional[tuple[str, float]]:
        FUZZY_PATTERNS = [
            ("salary payment staff",          "Salaries & Wages",       0.88),
            ("rent office premises",           "Rent",                   0.86),
            ("electricity bill payment",       "Electricity & Utilities",0.86),
            ("gst tax payment government",     "GST Payment",            0.90),
            ("atm cash withdrawal bank",       "ATM Cash Withdrawal",    0.90),
            ("professional consultant fee",    "Professional Fees",      0.84),
            ("insurance premium payment",      "Insurance Premium",      0.86),
            ("emi loan repayment bank",        "Loan Repayment",         0.88),
            ("purchase supplier vendor goods", "Purchase/Materials",     0.82),
            ("sales customer invoice payment", "Sales Revenue",          0.82),
        ]
        best_score  = 0
        best_result = None
        for pattern, account, base_conf in FUZZY_PATTERNS:
            score = fuzz.partial_ratio(narration.upper(), pattern.upper())
            if score > 70 and score > best_score:
                best_score  = score
                best_result = (account, base_conf * (score / 100))
        return best_result

    def classify_batch(
        self, transactions: list[ParsedTransaction]
    ) -> list[ClassifiedTransaction]:
        """Classify a list of ParsedTransaction objects."""
        result = []
        for i, t in enumerate(transactions):
            account, confidence = self.classify(t.narration)
            # Lower confidence for very large amounts (more likely to need review)
            if t.amount > 100000:
                confidence = max(confidence - 0.05, 0.55)
            result.append(ClassifiedTransaction(
                id          = f"bt-{i}-{abs(hash(t.narration + str(t.amount))) % 99999:05d}",
                txn_date    = t.txn_date,
                narration   = t.narration,
                amount      = t.amount,
                txn_type    = t.txn_type,
                balance     = t.balance,
                reference   = t.reference,
                account     = account,
                confidence  = round(confidence, 3),
                method      = "rule",
                status      = "unmatched",
            ))
        return result


# ═══════════════════════════════════════════════════════════════════════════════
# 3. INVOICE PARSER  (100% in-house — no OCR API needed)
# ═══════════════════════════════════════════════════════════════════════════════

class InvoiceParser:
    """
    Extracts invoice fields from PDF files using pdfplumber + regex.
    Handles: GST Tax Invoices, Purchase Bills, Receipts, Debit/Credit Notes.
    NO external API. Works on any text-based PDF.
    """

    PATTERNS = {
        "invoice_no": [
            r"invoice\s*(?:no|number|#)[:\s#]+([A-Z0-9/\-]+)",
            r"inv[:\s#]+([A-Z0-9/\-]+)",
            r"bill\s*no[:\s]+([A-Z0-9/\-]+)",
            r"ref(?:erence)?\s*no[:\s]+([A-Z0-9/\-]+)",
            r"receipt\s*no[:\s]+([A-Z0-9/\-]+)",
            r"voucher\s*no[:\s]+([A-Z0-9/\-]+)",
        ],
        "party_name": [
            r"bill(?:ed)?\s*to[:\s]+([^\n,]{3,60})",
            r"sold\s*to[:\s]+([^\n,]{3,60})",
            r"customer[:\s]+([^\n,]{3,60})",
            r"client[:\s]+([^\n,]{3,60})",
            r"party[:\s]+([^\n,]{3,60})",
            r"m/s[:\s]+([^\n,]{3,60})",
            r"buyer[:\s]+([^\n,]{3,60})",
        ],
        "party_gstin": [
            r"(?:buyer|customer|party)\s*gstin[:\s]+([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z])",
            r"gstin\s*(?:of\s*buyer|of\s*customer)?[:\s]+([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z])",
        ],
        "hsn_sac": [
            r"hsn[/\s]*sac[:\s]+(\d{4,8})",
            r"hsn\s*code[:\s]+(\d{4,8})",
            r"sac\s*code[:\s]+(\d{4,8})",
        ],
        "date": [
            r"invoice\s*date[:\s]+(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})",
            r"date\s*of\s*invoice[:\s]+(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})",
            r"bill\s*date[:\s]+(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})",
            r"date[:\s]+(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4})",
        ],
        "amount": [
            r"(?:taxable|net|sub)\s*(?:amount|value)[:\s₹]+([0-9,]+\.?\d*)",
            r"(?:basic|base)\s*(?:amount|value)[:\s₹]+([0-9,]+\.?\d*)",
            r"amount\s*before\s*tax[:\s₹]+([0-9,]+\.?\d*)",
        ],
        "cgst": [
            r"cgst\s*(?:@\s*[\d\.]+%)?[:\s₹]+([0-9,]+\.?\d*)",
            r"c\.gst[:\s₹]+([0-9,]+\.?\d*)",
        ],
        "sgst": [
            r"sgst\s*(?:@\s*[\d\.]+%)?[:\s₹]+([0-9,]+\.?\d*)",
            r"s\.gst[:\s₹]+([0-9,]+\.?\d*)",
        ],
        "igst": [
            r"igst\s*(?:@\s*[\d\.]+%)?[:\s₹]+([0-9,]+\.?\d*)",
            r"i\.gst[:\s₹]+([0-9,]+\.?\d*)",
        ],
        "total": [
            r"grand\s*total[:\s₹]+([0-9,]+\.?\d*)",
            r"total\s*amount[:\s₹]+([0-9,]+\.?\d*)",
            r"amount\s*due[:\s₹]+([0-9,]+\.?\d*)",
            r"net\s*payable[:\s₹]+([0-9,]+\.?\d*)",
            r"invoice\s*total[:\s₹]+([0-9,]+\.?\d*)",
            r"total[:\s₹]+([0-9,]+\.?\d*)",
        ],
    }

    def parse_pdf(self, content: bytes) -> ParsedInvoice:
        """Extract invoice fields from PDF bytes using pdfplumber."""
        if not PDF_AVAILABLE:
            raise RuntimeError("pdfplumber not installed.")

        text = ""
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"

        return self._extract_from_text(text)

    def parse_csv_excel(self, content: bytes, ext: str) -> ParsedInvoice:
        """Extract invoice data from structured Excel/CSV template."""
        if not EXCEL_AVAILABLE:
            raise RuntimeError("openpyxl not installed.")
        if ext == "csv":
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            data = {}
            for row in reader:
                if len(row) >= 2:
                    key = str(row[0]).lower().strip().replace(" ", "_")
                    data[key] = str(row[1]).strip()
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            data = {}
            for row in ws.iter_rows(values_only=True, max_row=40):
                if row[0] and row[1]:
                    key = str(row[0]).lower().strip().replace(" ", "_")
                    data[key] = str(row[1]).strip()

        def get_float(keys):
            for k in keys:
                v = data.get(k, "")
                try:
                    return float(str(v).replace(",", "").replace("₹", "").strip())
                except (ValueError, AttributeError):
                    pass
            return 0.0

        inv = ParsedInvoice()
        inv.invoice_no   = str(data.get("invoice_no", data.get("invoice_number", "")))
        inv.party_name   = str(data.get("party_name", data.get("customer_name", data.get("buyer", ""))))
        inv.party_gstin  = str(data.get("party_gstin", data.get("customer_gstin", "")))
        inv.hsn_sac      = str(data.get("hsn_sac", data.get("hsn", data.get("sac", ""))))
        inv.invoice_date = str(data.get("date", data.get("invoice_date", date.today().isoformat())))
        inv.amount       = get_float(["subtotal", "net_amount", "taxable_value", "amount"])
        inv.cgst         = get_float(["cgst"])
        inv.sgst         = get_float(["sgst"])
        inv.igst         = get_float(["igst"])
        inv.total        = get_float(["total", "grand_total", "amount_due"])
        if inv.total == 0 and inv.amount > 0:
            inv.total = inv.amount + inv.cgst + inv.sgst + inv.igst
        return inv

    def _extract_from_text(self, text: str) -> ParsedInvoice:
        """Extract invoice fields from raw text using regex."""
        result = ParsedInvoice()

        text_lower = text.lower()
        if any(w in text_lower for w in ["purchase order", "vendor", "supplier",
                                          "bill from", "received from"]):
            result.inv_type = "purchase"
        elif any(w in text_lower for w in ["bill to", "sold to", "customer",
                                             "client", "invoice to", "tax invoice"]):
            result.inv_type = "sales"

        def find(field: str) -> str:
            for pattern in self.PATTERNS.get(field, []):
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    return m.group(1).strip()
            return ""

        def find_amount(field: str) -> float:
            raw = find(field)
            if not raw:
                return 0.0
            try:
                return float(raw.replace(",", ""))
            except ValueError:
                return 0.0

        result.invoice_no   = find("invoice_no")[:50]
        result.party_name   = find("party_name")[:100]
        result.party_gstin  = find("party_gstin")[:20]
        result.hsn_sac      = find("hsn_sac")[:10]
        result.invoice_date = self._parse_date_str(find("date"))
        result.narration    = self._extract_narration(text)
        result.amount       = find_amount("amount")
        result.cgst         = find_amount("cgst")
        result.sgst         = find_amount("sgst")
        result.igst         = find_amount("igst")
        result.total        = find_amount("total")

        if result.total == 0 and result.amount > 0:
            result.total = result.amount + result.cgst + result.sgst + result.igst
        if result.amount == 0 and result.total > 0:
            result.amount = result.total - result.cgst - result.sgst - result.igst

        return result

    def _parse_date_str(self, raw: str) -> str:
        if not raw:
            return date.today().isoformat()
        parser = BankStatementParser()
        result = parser._parse_date(raw)
        return result or date.today().isoformat()

    def _extract_narration(self, text: str) -> str:
        patterns = [
            r"description\s*of\s*(?:goods|services)[:\s]+([^\n]+)",
            r"particulars[:\s]+([^\n]+)",
            r"goods/services[:\s]+([^\n]+)",
            r"item\s*description[:\s]+([^\n]+)",
        ]
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                return m.group(1).strip()[:150]
        return "Invoice payment"


# ═══════════════════════════════════════════════════════════════════════════════
# 4. VOUCHER AUTO-DETECTOR  (100% in-house)
# ═══════════════════════════════════════════════════════════════════════════════

class VoucherAutoDetector:
    """
    Automatically determines voucher type and suggests accounting entries
    from a narration/description — NO external API.

    Voucher types:
      - receipt:   money received from customer
      - payment:   money paid to vendor/expense
      - contra:    cash↔bank transfer, ATM withdrawal
      - journal:   adjustments, depreciation, provisions
      - sales:     sales invoice posting
      - purchase:  purchase invoice posting
      - credit_note:  sales return
      - debit_note:   purchase return
    """

    VOUCHER_RULES = [
        # Contra (bank↔cash)
        (["atm wdl", "atm cash", "cash withdrawal", "cash deposit",
          "transfer to self", "own account", "contra", "cash to bank",
          "bank to cash"],
         "contra", 0.97),

        # Sales voucher
        (["sales", "revenue", "income", "invoice payment",
          "payment received", "upi received", "amount received from customer"],
         "receipt", 0.92),

        # Purchase / Expense payment
        (["purchase", "material", "supplier", "vendor", "bill paid",
          "expense paid", "payment to", "paid to"],
         "payment", 0.90),

        # GST / Tax payment
        (["gst payment", "tds payment", "advance tax", "income tax",
          "gst challan", "tax challan"],
         "payment", 0.95),

        # Salary payment
        (["salary", "wages", "payroll", "staff payment"],
         "payment", 0.94),

        # Journal / Adjustment
        (["depreciation", "amortization", "provision", "adjustment",
          "write off", "accrual", "reversal entry", "opening balance",
          "closing entry", "transfer entry"],
         "journal", 0.92),

        # Credit note
        (["credit note", "sales return", "return inward", "goods returned by customer"],
         "credit_note", 0.95),

        # Debit note
        (["debit note", "purchase return", "return outward", "goods returned to supplier"],
         "debit_note", 0.95),
    ]

    # Suggested debit/credit accounts per voucher type
    ACCOUNT_SUGGESTIONS = {
        "receipt": {
            "debit":  "Bank Account / Cash",
            "credit": "Sales Revenue / Customer Account",
        },
        "payment": {
            "debit":  "Expense Account / Vendor Account",
            "credit": "Bank Account / Cash",
        },
        "contra": {
            "debit":  "Cash Account",
            "credit": "Bank Account",
        },
        "journal": {
            "debit":  "Expense / Asset Account",
            "credit": "Liability / Income Account",
        },
        "sales": {
            "debit":  "Customer / Debtor Account",
            "credit": "Sales Account + GST Payable",
        },
        "purchase": {
            "debit":  "Purchase Account + GST Input",
            "credit": "Supplier / Creditor Account",
        },
        "credit_note": {
            "debit":  "Sales Return Account",
            "credit": "Customer Account",
        },
        "debit_note": {
            "debit":  "Supplier Account",
            "credit": "Purchase Return Account",
        },
    }

    def detect(self, narration: str, amount: float = 0.0,
               txn_type: str = "") -> ParsedVoucher:
        """
        Detect voucher type from narration and transaction type.
        Returns ParsedVoucher with suggested accounts.
        """
        n = narration.upper().strip()

        # Use txn_type hint
        if txn_type == "credit":
            v_type, conf = "receipt", 0.80
        elif txn_type == "debit":
            v_type, conf = "payment", 0.80
        else:
            v_type, conf = "journal", 0.65

        # Override with keyword rules
        for keywords, vtype, confidence in self.VOUCHER_RULES:
            for kw in keywords:
                if kw.upper() in n:
                    v_type = vtype
                    conf   = confidence
                    break

        suggestion = self.ACCOUNT_SUGGESTIONS.get(v_type, {})

        return ParsedVoucher(
            voucher_type   = v_type,
            narration      = narration,
            amount         = amount,
            debit_account  = suggestion.get("debit",  ""),
            credit_account = suggestion.get("credit", ""),
            confidence     = conf,
        )

    def detect_batch(
        self, transactions: list[ParsedTransaction]
    ) -> list[ParsedVoucher]:
        return [
            self.detect(t.narration, t.amount, t.txn_type)
            for t in transactions
        ]


# ═══════════════════════════════════════════════════════════════════════════════
# 5. CONVENIENCE FUNCTIONS  (used by main.py endpoints)
# ═══════════════════════════════════════════════════════════════════════════════

# Module-level singletons (instantiated once per worker)
_parser     = BankStatementParser()
_classifier = TransactionClassifier()
_invoice    = InvoiceParser()
_voucher    = VoucherAutoDetector()


def parse_and_classify_statement(content: bytes, filename: str) -> dict:
    """
    Full in-house pipeline: parse bank statement → classify → detect voucher type.
    Returns dict with 'transactions' list and metadata.
    NO external API calls.
    """
    raw_txns = _parser.parse(content, filename)

    if not raw_txns:
        raise ValueError(
            f"No transactions found in '{filename}'. "
            "Make sure the file is a valid bank statement PDF/CSV/Excel. "
            "If it's a scanned (image) PDF, please export as CSV from your bank instead."
        )

    classified = _classifier.classify_batch(raw_txns)
    vouchers   = _voucher.detect_batch(raw_txns)

    result = []
    for t, v in zip(classified, vouchers):
        result.append({
            "id":                   t.id,
            "txn_date":             t.txn_date,
            "narration":            t.narration,
            "amount":               t.amount,
            "txn_type":             t.txn_type,
            "balance":              t.balance,
            "reference":            t.reference,
            "ai_suggested_account": t.account,
            "voucher_type":         v.voucher_type,
            "debit_account":        v.debit_account,
            "credit_account":       v.credit_account,
            "confidence":           t.confidence,
            "status":               "unmatched",
        })

    return {
        "transactions": result,
        "total_parsed": len(result),
        "filename":     filename,
        "file_type":    filename.rsplit(".", 1)[-1].lower() if "." in filename else "unknown",
    }


def parse_invoice_file(content: bytes, filename: str) -> dict:
    """
    Parse an invoice PDF/Excel/CSV and return structured fields.
    NO external API calls.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext == "pdf":
        inv = _invoice.parse_pdf(content)
    elif ext in ("xlsx", "xls", "csv"):
        inv = _invoice.parse_csv_excel(content, ext)
    else:
        inv = ParsedInvoice()
        inv.narration = "Please fill in the invoice details manually."

    return {
        "type":       inv.inv_type,
        "date":       inv.invoice_date,
        "reference":  inv.invoice_no,
        "party":      inv.party_name,
        "gstin":      inv.party_gstin,
        "hsn_sac":    inv.hsn_sac,
        "narration":  inv.narration,
        "amount":     inv.amount,
        "cgst":       inv.cgst,
        "sgst":       inv.sgst,
        "igst":       inv.igst,
        "total":      inv.total,
    }


def detect_voucher_type(narration: str, amount: float = 0.0,
                         txn_type: str = "") -> dict:
    """
    Detect voucher type and suggest journal entries for a narration.
    Used by Journal.jsx to auto-fill voucher form.
    NO external API calls.
    """
    v = _voucher.detect(narration, amount, txn_type)
    return {
        "voucher_type":   v.voucher_type,
        "narration":      v.narration,
        "amount":         v.amount,
        "debit_account":  v.debit_account,
        "credit_account": v.credit_account,
        "confidence":     v.confidence,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 6. VECTOR MEMORY  (persistent narration → account mapping)
# Added in v2 upgrade — supplements rule-based classifier above
# ═══════════════════════════════════════════════════════════════════════════════

import json as _json
import os as _os
from pathlib import Path as _Path

MEMORY_PATH = _Path(_os.getenv("AI_MEMORY_PATH", "/tmp/ai_memory.json"))

_NUMPY_AVAILABLE = False
try:
    import numpy as _np
    _NUMPY_AVAILABLE = True
except ImportError:
    pass

_EMBED_AVAILABLE = False
_embed_model = None
try:
    from sentence_transformers import SentenceTransformer as _ST
    _embed_model = _ST(_os.getenv("EMBEDDING_MODEL", "all-MiniLM-L6-v2"))
    _EMBED_AVAILABLE = True
except Exception:
    pass


class VectorMemory:
    """
    Flat vector store: {narration → (account_code, embedding)}.
    Persisted as JSON. For production, replace with pgvector or Chroma.
    Works alongside TransactionClassifier — called first for user-corrected entries.
    """

    def __init__(self):
        self._store: list[dict] = []
        self._load()

    def _load(self):
        if MEMORY_PATH.exists():
            try:
                self._store = _json.loads(MEMORY_PATH.read_text())
            except Exception:
                pass

    def _save(self):
        try:
            MEMORY_PATH.write_text(_json.dumps(self._store))
        except Exception:
            pass

    def _embed(self, text: str):
        if not _EMBED_AVAILABLE or _embed_model is None:
            return None
        return _embed_model.encode(text, normalize_embeddings=True).tolist()

    def add(self, narration: str, account_code: str):
        emb = self._embed(narration)
        self._store = [s for s in self._store if s["narration"] != narration]
        self._store.append({"narration": narration, "account_code": account_code, "embedding": emb})
        self._save()

    def search(self, query: str, top_k: int = 3) -> list[dict]:
        if not _EMBED_AVAILABLE or not _NUMPY_AVAILABLE or not self._store:
            return []
        q_vec = self._embed(query)
        if q_vec is None:
            return []
        q_arr = _np.array(q_vec)
        results = []
        for entry in self._store:
            if entry.get("embedding") is None:
                continue
            score = float(_np.dot(q_arr, _np.array(entry["embedding"])))
            results.append({"narration": entry["narration"], "account_code": entry["account_code"], "score": score})
        results.sort(key=lambda x: x["score"], reverse=True)
        return results[:top_k]

    def best_match(self, query: str, threshold: float = 0.80):
        hits = self.search(query, top_k=1)
        return hits[0] if hits and hits[0]["score"] >= threshold else None

    def size(self) -> int:
        return len(self._store)


_memory_singleton = None

def get_memory() -> VectorMemory:
    global _memory_singleton
    if _memory_singleton is None:
        _memory_singleton = VectorMemory()
    return _memory_singleton
